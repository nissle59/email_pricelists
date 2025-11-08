from datetime import datetime

import pandas as pd
import re
import json

from openpyxl.styles import Side, Border, PatternFill, Font
from openpyxl.utils import get_column_letter
from openpyxl import load_workbook

from models import ParsingConfig
from utils.paths import pm


def old_to_excel_with_role_widths(df: pd.DataFrame, filename: str, widths: dict | None = None):
    """
    Сохраняет DataFrame в Excel и устанавливает ширину колонок по ролям.

    Аргументы:
    - df: DataFrame
    - filename: путь к Excel файлу
    - widths: словарь {роль: ширина}
    """
    role_widths = {
        "Наименование": 100,
        "Артикул": 18,
        "Цена": 16,
        "Остаток": 16,
        "Бренд": 18,
        "РРЦ": 16
    }
    if not widths:
        widths = role_widths
    df.to_excel(filename, index=False)
    wb = load_workbook(filename)
    ws = wb.active

    for i, col in enumerate(df.columns, 1):
        width = widths.get(col, 15)  # если роль не в словаре, ставим ширину по умолчанию
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(filename)


def to_excel_with_role_widths(df: pd.DataFrame, filename: str, widths: dict | None = None):
    """
    Сохраняет DataFrame в Excel и устанавливает ширину колонок по ролям.

    Аргументы:
    - df: DataFrame
    - filename: путь к Excel файлу
    - widths: словарь {роль: ширина}
    """
    role_widths = {
        "Наименование": 100,
        "Артикул": 18,
        "Цена": 16,
        "Остаток": 16,
        "(?) Бренд": 18,
        "(?) РРЦ": 16
    }
    if not widths:
        widths = role_widths

    # Если есть колонка "Поставщик", перемещаем ее в конец
    if "Поставщик" in df.columns:
        cols = [col for col in df.columns if col != "Поставщик"] + ["Поставщик"]
        df = df[cols]

    # Сохраняем DataFrame в Excel
    df.to_excel(filename, index=False)

    # Открываем книгу для форматирования
    wb = load_workbook(filename)
    ws = wb.active

    # Устанавливаем ширину колонок
    for i, col in enumerate(df.columns, 1):
        width = widths.get(col, 15)  # если роль не в словаре, ставим ширину по умолчанию
        ws.column_dimensions[get_column_letter(i)].width = width

    # Добавляем автофильтр на заголовки
    if len(df) > 0:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(df.columns))}{len(df) + 1}"

    # Форматируем заголовки (цвет фона и границы)
    header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Применяем форматирование к заголовкам
    for cell in ws[1]:
        cell.fill = header_fill
        cell.border = thin_border
        cell.font = Font(bold=True)

    # Применяем границы ко всем ячейкам с данными
    if len(df) > 0:
        for row in ws.iter_rows(min_row=1, max_row=len(df) + 1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border

    # Сохраняем изменения
    wb.save(filename)



def apply_parser_settings(df_original: pd.DataFrame, settings: ParsingConfig, vendor_name: str,
                          date: datetime | None = None, quantum_config: dict | None = None) -> pd.DataFrame:
    """
    Применяет настройки парсинга к исходному DataFrame и возвращает отфильтрованный DataFrame.

    Параметры:
    - df_original: исходный DataFrame (все строки Excel)
    - settings_file: путь к JSON-файлу с настройками

    Возвращает:
    - df_filtered: DataFrame с колонками, переименованными в роли, только валидные записи
    """
    # with open(settings_file, "r", encoding="utf-8") as f:
    #     settings = json.load(f)
    def calculate_quantum_value(row):
        """Вычисляет значение кванта на основе конфигурации"""
        if not quantum_config:
            return None

        try:
            quantum_col = quantum_config['quantum_column']
            if not quantum_col or quantum_col not in row:
                return None

            unit = str(row[quantum_col]).strip().lower()
            if not unit:
                return None

            # Ищем сопоставление для единицы измерения
            for unit_pattern, target_column in quantum_config['unit_mappings'].items():
                if unit_pattern.lower() in unit or unit in unit_pattern.lower():
                    if target_column == "1":
                        return 1
                    elif target_column in row and pd.notna(row[target_column]):
                        try:
                            return float(row[target_column])
                        except (ValueError, TypeError):
                            pass

                    # Проверяем специальные колонки
                    if (unit_pattern.lower() in ['кор', 'коробка', 'кор.'] and
                            quantum_config['box_quantity_column'] in row):
                        try:
                            return float(row[quantum_config['box_quantity_column']])
                        except (ValueError, TypeError):
                            pass

                    if (unit_pattern.lower() in ['бл', 'блок', 'дисплейбокс'] and
                            quantum_config['block_quantity_column'] in row):
                        try:
                            return float(row[quantum_config['block_quantity_column']])
                        except (ValueError, TypeError):
                            pass

            # Если не нашли сопоставление, пробуем распарсить как число
            try:
                return float(unit)
            except (ValueError, TypeError):
                return 1  # Значение по умолчанию

        except Exception as e:
            print(f"Ошибка вычисления кванта: {e}")
            return 1
    if not settings.active:
        return pd.DataFrame([])

    header_row = settings.header_row
    roles_mapping = {}
    for mapping in settings.mappings:
        roles_mapping.update({mapping.role.name: mapping.column_name})
    # roles_mapping = settings.roles_mapping  # dict: роль -> исходная колонка

    # формируем DataFrame с правильными колонками
    df = df_original.copy()
    # объединяем шапку, если она многострочная
    header_part = df.iloc[header_row:header_row + 1]
    headers = header_part.fillna("").astype(str).agg(lambda col: " ".join([v.strip() for v in col if v.strip()]),
                                                     axis=0)
    df = df.iloc[header_row + 1:].copy()
    df.columns = headers
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # отбрасываем колонки без названия
    valid_columns = [c for c in df.columns if c.strip() != ""]
    df = df[valid_columns]

    # оставляем только колонки из ролей, которые есть в DataFrame
    used_columns = [roles_mapping[r] for r in roles_mapping if roles_mapping[r] in df.columns]
    df_filtered = df[used_columns].copy()

    # фильтрация строк ---------------------------------
    # имя и цена по ролям
    name_role = None
    price_role = None
    for r, col in roles_mapping.items():
        if r.lower() in ["наименование"]:
            name_role = r
        elif r.lower() in ["закупочная цена"]:
            price_role = r

    quantum_col = roles_mapping.get("Квант")

    if quantum_col and quantum_config and quantum_col == quantum_config['quantum_column']:
        # Применяем сложную логику для кванта
        print("Применяем сложную логику для кванта...")
        df_filtered[quantum_col] = df.apply(calculate_quantum_value, axis=1)
    elif quantum_col:
        # Стандартная обработка кванта
        try:
            df_filtered[quantum_col] = pd.to_numeric(df_filtered[quantum_col], errors='coerce')
        except:
            df_filtered[quantum_col] = 1



    # переименуем колонки на роли
    df_filtered.columns = [r for r in roles_mapping if roles_mapping[r] in df_filtered.columns]

    # фильтруем по названию
    if name_role and name_role in df_filtered.columns:
        df_filtered = df_filtered[
            df_filtered[name_role].notna() & (df_filtered[name_role].astype(str).str.strip() != "")]

    # фильтруем по цене
    def valid_price(v):
        if pd.isna(v):
            return False
        if isinstance(v, (int, float)):
            return True
        s = str(v).strip()
        match = re.search(r"\d+(\.\d+)?", s)
        if not match:
            return False
        return len(s) <= 10  # ограничение по длине

    if price_role and price_role in df_filtered.columns:
        df_filtered = df_filtered[df_filtered[price_role].apply(valid_price)]

    df_filtered["Поставщик"] = vendor_name

    if date is not None:
        df_filtered["Дата"] = date

    if settings.save_parsed:
        out_fname = f"{vendor_name} - {settings.name} - {date.strftime('%d.%m.%Y %H-%M')}.xlsx"
        to_excel_with_role_widths(df_filtered.drop(["Дата"], axis=1), pm.save_file(out_fname, mode='parsed'))

    if not settings.to_common:
        return pd.DataFrame([])

    return df_filtered
